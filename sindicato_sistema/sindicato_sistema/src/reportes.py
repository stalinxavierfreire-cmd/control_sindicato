import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from config import PDF_REPORTS_DIR, EXCEL_REPORTS_DIR
from src.asociados import GestionAsociados
from src.comprobantes import GestionComprobantes
from src.sistema import SistemaSindical
from src.decorators import handle_errors

class GeneradorReportes:
    @staticmethod
    @handle_errors
    def generar_reporte_asociados_pdf(filtros: Dict[str, Any] = None) -> str:
        if filtros is None:
            filtros = {}
        asociados = GestionAsociados.buscar_asociados(**filtros)
        nombre_sindicato = SistemaSindical().obtener_configuracion("nombre_sindicato")
        fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
        nombre_archivo = os.path.join(
            PDF_REPORTS_DIR,
            f"reporte_asociados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        doc = SimpleDocTemplate(nombre_archivo, pagesize=letter)
        elementos = []
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle(
            "titulo", parent=estilos["Heading1"], fontName="Helvetica-Bold", alignment=1, spaceAfter=12
        )
        estilo_subtitulo = ParagraphStyle(
            "subtitulo", parent=estilos["Heading2"], fontName="Helvetica", alignment=1, spaceAfter=12
        )
        estilo_normal = ParagraphStyle(
            "normal", parent=estilos["Normal"], fontName="Helvetica", spaceAfter=6
        )
        elementos.append(Paragraph(nombre_sindicato, estilo_titulo))
        elementos.append(Paragraph("Reporte de Asociados", estilo_subtitulo))
        elementos.append(Paragraph(f"Fecha: {fecha_reporte}", estilo_normal))
        if filtros:
            texto_filtros = "Filtros aplicados: " + ", ".join(
                f"{k}: {v}" for k, v in filtros.items() if v
            )
            elementos.append(Paragraph(texto_filtros, estilo_normal))
        elementos.append(Spacer(1, 0.25 * inch))
        if asociados:
            datos = [
                ["ID", "Nombre", "Apellido", "DNI", "Teléfono", "Estado", "Ingreso"]
            ]
            for a in asociados:
                datos.append([
                    str(a["id"]),
                    a["nombre"],
                    a["apellido"],
                    a["dni"],
                    a["telefono"] or "-",
                    a["estado"].capitalize(),
                    datetime.strptime(a["fecha_ingreso"], "%Y-%m-%d").strftime("%d/%m/%Y")
                ])
            tabla = Table(datos)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]))
            elementos.append(tabla)
        else:
            elementos.append(Paragraph("No se encontraron asociados con los criterios especificados.", estilo_normal))
        doc.build(elementos)
        return nombre_archivo

    @staticmethod
    @handle_errors
    def generar_reporte_financiero_pdf(filtros: Dict[str, Any] = None) -> str:
        if filtros is None:
            filtros = {}
        comprobantes = GestionComprobantes.obtener_comprobantes(**filtros)
        resumen = GestionComprobantes.obtener_resumen_financiero()
        nombre_sindicato = SistemaSindical().obtener_configuracion("nombre_sindicato")
        fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
        nombre_archivo = os.path.join(
            PDF_REPORTS_DIR,
            f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        doc = SimpleDocTemplate(nombre_archivo, pagesize=letter)
        elementos = []
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle(
            "titulo", parent=estilos["Heading1"], fontName="Helvetica-Bold", alignment=1, spaceAfter=12
        )
        estilo_subtitulo = ParagraphStyle(
            "subtitulo", parent=estilos["Heading2"], fontName="Helvetica", alignment=1, spaceAfter=12
        )
        estilo_normal = ParagraphStyle(
            "normal", parent=estilos["Normal"], fontName="Helvetica", spaceAfter=6
        )
        estilo_resumen = ParagraphStyle(
            "resumen", parent=estilos["Normal"], fontName="Helvetica-Bold", spaceAfter=12
        )
        elementos.append(Paragraph(nombre_sindicato, estilo_titulo))
        elementos.append(Paragraph("Reporte Financiero", estilo_subtitulo))
        elementos.append(Paragraph(f"Fecha: {fecha_reporte}", estilo_normal))
        if filtros:
            texto_filtros = "Filtros aplicados: " + ", ".join(
                f"{k}: {v}" for k, v in filtros.items() if v
            )
            elementos.append(Paragraph(texto_filtros, estilo_normal))
        elementos.append(Spacer(1, 0.25 * inch))
        elementos.append(Paragraph("Resumen Financiero", estilo_resumen))
        datos_resumen = [
            ["Total Ingresos", f"${resumen['ingresos']:,.2f}"],
            ["Total Egresos", f"${resumen['egresos']:,.2f}"],
            ["Balance", f"${resumen['balance']:,.2f}"]
        ]
        tabla_resumen = Table(datos_resumen, colWidths=[200, 100])
        tabla_resumen.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elementos.append(tabla_resumen)
        elementos.append(Spacer(1, 0.25 * inch))
        if comprobantes:
            elementos.append(Paragraph("Detalle de Comprobantes", estilo_resumen))
            datos = [
                ["Fecha", "Tipo", "Monto", "Descripción", "Asociado"]
            ]
            for c in comprobantes:
                datos.append([
                    datetime.strptime(c["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y"),
                    "Ingreso" if c["tipo"] == "ingreso" else "Egreso",
                    f"${c['monto']:,.2f}",
                    c["descripcion"],
                    c["asociado_nombre"] or "-"
                ])
            tabla = Table(datos)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]))
            elementos.append(tabla)
        else:
            elementos.append(Paragraph("No se encontraron comprobantes con los criterios especificados.", estilo_normal))
        doc.build(elementos)
        return nombre_archivo

    @staticmethod
    @handle_errors
    def generar_reporte_asociados_excel(filtros: Dict[str, Any] = None) -> str:
        if filtros is None:
            filtros = {}
        asociados = GestionAsociados.buscar_asociados(**filtros)
        nombre_sindicato = SistemaSindical().obtener_configuracion("nombre_sindicato")
        fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
        nombre_archivo = os.path.join(
            EXCEL_REPORTS_DIR,
            f"reporte_asociados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asociados"
        ws.merge_cells("A1:G1")
        ws["A1"] = nombre_sindicato
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:G2")
        ws["A2"] = "Reporte de Asociados"
        ws["A2"].font = Font(bold=True, size=12)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:G3")
        ws["A3"] = f"Fecha: {fecha_reporte}"
        ws["A3"].alignment = Alignment(horizontal="center")
        if filtros:
            ws.merge_cells("A4:G4")
            texto_filtros = "Filtros aplicados: " + ", ".join(
                f"{k}: {v}" for k, v in filtros.items() if v
            )
            ws["A4"] = texto_filtros
            ws["A4"].alignment = Alignment(horizontal="left")
            fila_inicio = 6
        else:
            fila_inicio = 5
        encabezados = ["ID", "Nombre", "Apellido", "DNI", "Teléfono", "Estado", "Fecha Ingreso"]
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="left")
        for fila, asociado in enumerate(asociados, start=fila_inicio+1):
            ws.cell(row=fila, column=1, value=asociado["id"])
            ws.cell(row=fila, column=2, value=asociado["nombre"])
            ws.cell(row=fila, column=3, value=asociado["apellido"])
            ws.cell(row=fila, column=4, value=asociado["dni"])
            ws.cell(row=fila, column=5, value=asociado["telefono"] or "-")
            ws.cell(row=fila, column=6, value=asociado["estado"].capitalize())
            ws.cell(
                row=fila, column=7, 
                value=datetime.strptime(asociado["fecha_ingreso"], "%Y-%m-%d").strftime("%d/%m/%Y")
            )
        for col in range(1, 8):
            col_letra = get_column_letter(col)
            if col == 1:
                ws.column_dimensions[col_letra].width = 8
            elif col in (2, 3):
                ws.column_dimensions[col_letra].width = 20
            elif col == 4:
                ws.column_dimensions[col_letra].width = 15
            elif col == 5:
                ws.column_dimensions[col_letra].width = 15
            elif col == 6:
                ws.column_dimensions[col_letra].width = 12
            elif col == 7:
                ws.column_dimensions[col_letra].width = 12
        wb.save(nombre_archivo)
        return nombre_archivo

    @staticmethod
    @handle_errors
    def generar_reporte_financiero_excel(filtros: Dict[str, Any] = None) -> str:
        if filtros is None:
            filtros = {}
        comprobantes = GestionComprobantes.obtener_comprobantes(**filtros)
        resumen = GestionComprobantes.obtener_resumen_financiero()
        nombre_sindicato = SistemaSindical().obtener_configuracion("nombre_sindicato")
        fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
        nombre_archivo = os.path.join(
            EXCEL_REPORTS_DIR,
            f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Finanzas"
        ws.merge_cells("A1:E1")
        ws["A1"] = nombre_sindicato
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:E2")
        ws["A2"] = "Reporte Financiero"
        ws["A2"].font = Font(bold=True, size=12)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:E3")
        ws["A3"] = f"Fecha: {fecha_reporte}"
        ws["A3"].alignment = Alignment(horizontal="center")
        if filtros:
            ws.merge_cells("A4:E4")
            texto_filtros = "Filtros aplicados: " + ", ".join(
                f"{k}: {v}" for k, v in filtros.items() if v
            )
            ws["A4"] = texto_filtros
            ws["A4"].alignment = Alignment(horizontal="left")
            fila_inicio = 6
        else:
            fila_inicio = 5
        ws.cell(row=fila_inicio, column=1, value="Resumen Financiero").font = Font(bold=True)
        fila_inicio += 1
        ws.cell(row=fila_inicio, column=1, value="Total Ingresos")
        ws.cell(row=fila_inicio, column=2, value=resumen["ingresos"]).number_format = '"$"#,##0.00'
        fila_inicio += 1
        ws.cell(row=fila_inicio, column=1, value="Total Egresos")
        ws.cell(row=fila_inicio, column=2, value=resumen["egresos"]).number_format = '"$"#,##0.00'
        fila_inicio += 1
        ws.cell(row=fila_inicio, column=1, value="Balance")
        ws.cell(row=fila_inicio, column=2, value=resumen["balance"]).number_format = '"$"#,##0.00'
        ws.cell(row=fila_inicio, column=2).font = Font(bold=True)
        fila_inicio += 2
        encabezados = ["Fecha", "Tipo", "Monto", "Descripción", "Asociado"]
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="left")
        fila_inicio += 1
        for fila, comprobante in enumerate(comprobantes, start=fila_inicio):
            ws.cell(
                row=fila, column=1, 
                value=datetime.strptime(comprobante["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y")
            )
            ws.cell(
                row=fila, column=2, 
                value="Ingreso" if comprobante["tipo"] == "ingreso" else "Egreso"
            )
            ws.cell(row=fila, column=3, value=comprobante["monto"]).number_format = '"$"#,##0.00'
            ws.cell(row=fila, column=4, value=comprobante["descripcion"])
            ws.cell(row=fila, column=5, value=comprobante["asociado_nombre"] or "-")
        for col in range(1, 6):
            col_letra = get_column_letter(col)
            if col == 1:
                ws.column_dimensions[col_letra].width = 12
            elif col == 2:
                ws.column_dimensions[col_letra].width = 10
            elif col == 3:
                ws.column_dimensions[col_letra].width = 12
            elif col == 4:
                ws.column_dimensions[col_letra].width = 40
            elif col == 5:
                ws.column_dimensions[col_letra].width = 25
        wb.save(nombre_archivo)
        return nombre_archivo